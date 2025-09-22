"""
data_load.py
Reads and validates DOCX and XLSX files for PPR dashboard.
"""
import os
import logging
from typing import Dict, Any
import pandas as pd
from openpyxl import load_workbook

# Setup logging
audit_log = []
logging.basicConfig(level=logging.INFO)

def read_markdown(file_path: str) -> str:
    try:
        with open(file_path, encoding='utf-8') as f:
            text = f.read()
        audit_log.append(f"Read Markdown: {file_path}")
        return text
    except Exception as e:
        audit_log.append(f"ERROR reading Markdown {file_path}: {e}")
        raise

def read_xlsx(file_path: str, sheet_name: str = None) -> pd.DataFrame:
    try:
        wb = load_workbook(file_path, data_only=True)
        if not sheet_name:
            sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        data = list(ws.values)
        columns = list(data[0])
        # Read all data without filtering - let validate functions handle column selection
        df = pd.DataFrame(data[1:], columns=columns)
        audit_log.append(f"Read XLSX: {file_path} [{sheet_name}] (all columns)")
        return df
    except Exception as e:
        audit_log.append(f"ERROR reading XLSX {file_path}: {e}")
        raise

def validate_national(df: pd.DataFrame) -> pd.DataFrame:
    # Accept 'Specie' and map to 'Species'
    if "Specie" in df.columns and "Species" not in df.columns:
        df = df.rename(columns={"Specie": "Species"})
        audit_log.append("Mapped 'Specie' to 'Species'.")
    # Use VADEMOS Forecasted Value for population
    if "VADEMOS Forecasted Value" in df.columns:
        df["Population"] = df["VADEMOS Forecasted Value"].fillna(0)
        audit_log.append("Set 'Population' from 'VADEMOS Forecasted Value'.")
    # Only keep required columns
    required_cols = ["Country", "Species", "Population", "Political_Stability_Index"]
    df = df[[col for col in required_cols if col in df.columns]]
    # Remove duplicate columns and columns with None as name
    df = df.loc[:,~df.columns.duplicated()]
    df = df.loc[:,df.columns.notnull()]
    return df

def validate_subregions(df: pd.DataFrame) -> pd.DataFrame:
    # Debug: Show what columns we actually have
    print(f"DEBUG - Subregions raw columns: {df.columns.tolist()}")
    print(f"DEBUG - Subregions sample data:\n{df.head()}")
    
    required_cols = ["Country", "Subregion", "Specie", "100%_Coverage", "api_name"]
    # Fill missing columns with default values
    for col in required_cols:
        if col not in df.columns:
            if col == "api_name":
                df[col] = ""  # Empty string for missing api_name
            elif col == "100%_Coverage":
                df[col] = 0
            else:
                df[col] = "Unknown"
            audit_log.append(f"Missing column '{col}' in Subregions.xlsx. Defaulted.")
        else:
            audit_log.append(f"Found column '{col}' in Subregions.xlsx.")
    # Only keep required columns
    df = df[required_cols]
    # Ensure 100%_Coverage is numeric
    df["100%_Coverage"] = pd.to_numeric(df["100%_Coverage"], errors='coerce').fillna(0)
    return df

def main():
    base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    files = {
        "methodology": os.path.join(base_path, "docs", "methodology.md"),
        "regional_costs": os.path.join(base_path, "docs", "regional_costs.md"),
        "country_case_costs": os.path.join(base_path, "docs", "country_case_costs.md"),
        "data_sources": os.path.join(base_path, "docs", "data_sources.md"),
        "national": os.path.join(base_path, "data", "National.xlsx"),
        "subregions": os.path.join(base_path, "data", "Subregions.xlsx"),
    }
    # Read Markdown files
    docs = {}
    for k in ["methodology", "regional_costs", "country_case_costs", "data_sources"]:
        docs[k] = read_markdown(files[k])
    # Read XLSX files
    national_df = validate_national(read_xlsx(files["national"]))
    subregions_df = validate_subregions(read_xlsx(files["subregions"]))
    # Log summary
    logging.info("Audit log:")
    for entry in audit_log:
        logging.info(entry)
    return {
        "docs": docs,
        "national_df": national_df,
        "subregions_df": subregions_df,
        "audit_log": audit_log,
    }

if __name__ == "__main__":
    main()
